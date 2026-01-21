"""
Data Export API
Handles CSV, JSON, Excel export with filtering and metadata
"""

from flask import Blueprint, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
import pandas as pd
import json
import io
from datetime import datetime
from io import BytesIO
import os

export_bp = Blueprint('export', __name__)

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'sensor_data.db')

def get_sensor_data(start_date=None, end_date=None, device_id=None, metrics=None):
    """Fetch sensor data with optional filtering"""
    conn = sqlite3.connect(DB_PATH)
    
    query = "SELECT id, timestamp, device_id, temperature, humidity, gas_analog, gas_digital FROM sensor_readings WHERE 1=1"
    params = []
    
    if start_date:
        query += " AND timestamp >= ?"
        params.append(start_date)
    if end_date:
        query += " AND timestamp <= ?"
        params.append(end_date)
    if device_id:
        query += " AND device_id = ?"
        params.append(device_id)
    
    query += " ORDER BY timestamp ASC"
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    
    if metrics:
        # Filter columns
        available_cols = ['id', 'timestamp', 'device_id', 'temperature', 'humidity', 'gas_analog', 'gas_digital']
        valid_metrics = [m for m in metrics if m in available_cols]
        if valid_metrics:
            df = df[valid_metrics]
    
    return df

@export_bp.route('/export/csv', methods=['GET', 'POST'])
def export_csv():
    """Export sensor data as CSV"""
    try:
        if request.method == 'POST':
            data = request.get_json()
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            device_id = data.get('device_id')
            metrics = data.get('metrics')
        else:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            device_id = request.args.get('device_id')
            metrics = request.args.getlist('metrics') if request.args.get('metrics') else None
        
        df = get_sensor_data(start_date, end_date, device_id, metrics)
        
        if df.empty:
            return jsonify({'error': 'No data found for the specified criteria'}), 404
        
        # Create CSV in memory
        output = io.StringIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        # Create response
        output_bytes = BytesIO()
        output_bytes.write(output.getvalue().encode('utf-8'))
        output_bytes.seek(0)
        
        filename = f"sensor_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return send_file(
            output_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@export_bp.route('/export/json', methods=['GET', 'POST'])
def export_json():
    """Export sensor data as JSON"""
    try:
        if request.method == 'POST':
            data = request.get_json()
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            device_id = data.get('device_id')
            metrics = data.get('metrics')
            include_metadata = data.get('include_metadata', True)
        else:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            device_id = request.args.get('device_id')
            metrics = request.args.getlist('metrics') if request.args.get('metrics') else None
            include_metadata = request.args.get('include_metadata', 'true').lower() == 'true'
        
        df = get_sensor_data(start_date, end_date, device_id, metrics)
        
        if df.empty:
            return jsonify({'error': 'No data found for the specified criteria'}), 404
        
        # Convert to JSON
        records = df.to_dict('records')
        
        result = {
            'data': records,
            'count': len(records)
        }
        
        if include_metadata:
            result['metadata'] = {
                'export_date': datetime.now().isoformat(),
                'start_date': start_date,
                'end_date': end_date,
                'device_id': device_id,
                'metrics': metrics,
                'total_records': len(records)
            }
            
            # Add statistics
            if 'temperature' in df.columns:
                result['metadata']['statistics'] = {
                    'temperature': {
                        'mean': float(df['temperature'].mean()),
                        'std': float(df['temperature'].std()),
                        'min': float(df['temperature'].min()),
                        'max': float(df['temperature'].max())
                    } if not df['temperature'].isna().all() else None
                }
            if 'humidity' in df.columns:
                if 'statistics' not in result['metadata']:
                    result['metadata']['statistics'] = {}
                result['metadata']['statistics']['humidity'] = {
                    'mean': float(df['humidity'].mean()),
                    'std': float(df['humidity'].std()),
                    'min': float(df['humidity'].min()),
                    'max': float(df['humidity'].max())
                } if not df['humidity'].isna().all() else None
        
        if request.method == 'POST':
            return jsonify(result), 200
        else:
            # For GET, send as file
            output = BytesIO()
            output.write(json.dumps(result, indent=2, default=str).encode('utf-8'))
            output.seek(0)
            
            filename = f"sensor_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            return send_file(
                output,
                mimetype='application/json',
                as_attachment=True,
                download_name=filename
            )
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@export_bp.route('/export/excel', methods=['POST'])
def export_excel():
    """Export sensor data as Excel with multiple sheets"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return jsonify({'error': 'openpyxl not installed. Install with: pip install openpyxl'}), 500
        
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        device_id = data.get('device_id')
        metrics = data.get('metrics')
        include_statistics = data.get('include_statistics', True)
        
        df = get_sensor_data(start_date, end_date, device_id, metrics)
        
        if df.empty:
            return jsonify({'error': 'No data found for the specified criteria'}), 404
        
        # Create Excel workbook in memory
        wb = Workbook()
        
        # Sheet 1: Data
        ws_data = wb.active
        ws_data.title = "Sensor Data"
        
        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws_data.cell(row=row_num, column=col_num, value=value)
        
        # Sheet 2: Statistics (if requested)
        if include_statistics:
            ws_stats = wb.create_sheet(title="Statistics")
            
            stats_data = []
            stats_data.append(['Metric', 'Mean', 'Std Dev', 'Min', 'Max', 'Count'])
            
            for col in ['temperature', 'humidity', 'gas_analog']:
                if col in df.columns and not df[col].isna().all():
                    stats_data.append([
                        col.capitalize(),
                        float(df[col].mean()),
                        float(df[col].std()),
                        float(df[col].min()),
                        float(df[col].max()),
                        int(df[col].count())
                    ])
            
            for row_num, row_data in enumerate(stats_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_stats.cell(row=row_num, column=col_num, value=value)
                    if row_num == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
        
        # Sheet 3: Metadata
        ws_meta = wb.create_sheet(title="Metadata")
        metadata = [
            ['Export Date', datetime.now().isoformat()],
            ['Start Date', start_date or 'All'],
            ['End Date', end_date or 'All'],
            ['Device ID', device_id or 'All'],
            ['Metrics', ', '.join(metrics) if metrics else 'All'],
            ['Total Records', len(df)]
        ]
        
        for row_num, (key, value) in enumerate(metadata, 1):
            ws_meta.cell(row=row_num, column=1, value=key).font = Font(bold=True)
            ws_meta.cell(row=row_num, column=2, value=value)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"sensor_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 400
